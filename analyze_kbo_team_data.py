import sys
sys.path.insert(0, '/Users/joshuakim/PycharmProjects/Math_Tools')

from matrix import *
from statistics_tools import *
import xlrd
import xlsxwriter
import openpyxl

from format_kbo_data import *


def format_rank_data():
    rank_history = dict()
    loc_dream = ""
    loc_magic = ""
    for year in range(1982, 2019):
        loc = "/Users/joshuakim/PycharmProjects/KBO_Analysis/kbo_team_rank_data_" + str(year) + ".xlsx"
        split = False
        try:
            workbook = openpyxl.load_workbook(loc)
        except:
            split = True
            loc_dream = "/Users/joshuakim/PycharmProjects/KBO_Analysis/kbo_team_rank_data_" + str(year) + "_dream.xlsx"
            loc_magic = "/Users/joshuakim/PycharmProjects/KBO_Analysis/kbo_team_rank_data_" + str(year) + "_magic.xlsx"
        if not split:
            workbook = openpyxl.load_workbook(loc)
            worksheet = workbook.active
            for row in range(2, worksheet.max_row+1):
                team_name = worksheet.cell(row=row, column=2).value.strip()
                if team_name not in rank_history:
                    rank_history[team_name] = []
                rank_history[team_name].append(worksheet.cell(row=row, column=7).value)
        else:
            for loc in [loc_dream, loc_magic]:
                workbook = openpyxl.load_workbook(loc)
                worksheet = workbook.active
                for row in range(2, worksheet.max_row + 1):
                    team_name = worksheet.cell(row=row, column=2).value.strip()
                    if team_name not in rank_history:
                        rank_history[team_name] = []
                    rank_history[team_name].append(worksheet.cell(row=row, column=7).value)
    return rank_history


def find_rank_average():
    rank_history = format_rank_data()
    rank_average = {key: sum(rank_history[key]) / len(rank_history[key]) for key in rank_history}
    return sorted(rank_average.items(), key=lambda kv: kv[1])


def analyze_rank_data_wrt_time(num_years):
    rank_history = format_rank_data()
    rank_input = []
    rank_response = []
    rank_history_formatted = {key: rank_history[key] for key in rank_history if len(rank_history[key]) > num_years}
    for i in range(num_years):
        rank_input.append([])
    for team in rank_history_formatted:
        for i in range(num_years):
            rank_input[i].extend(rank_history_formatted[team][i:i-num_years])
        rank_response.extend(rank_history_formatted[team][num_years:])
    return linear_least_square_fit(*rank_input, response=rank_response), multiple_correlation(*rank_input, response=rank_response)


def analyze_team_data(years, categories):
    rank_total, data_total = format_team_data(years, categories)

    # #########################
    # # ONLY FOR TESTING PURPOSE
    # print("Ranking: ", rank_total)
    # print("Data: ", data_total)
    # # REMOVE THIS LATER
    # #########################
    
    return linear_least_square_fit(*data_total, response=rank_total), multiple_correlation(*data_total, response=rank_total)


def compare_two_categories(years, categories):
    if len(categories) != 2:
        print("incorrect category data input")
        return False
    rank_total, data_total = format_team_data(years, categories)
    return linear_least_square_fit(data_total[0], response=data_total[1]), pearson_correlation_sample(data_total[0], data_total[1])


def format_team_data(years, categories):
    # category lists
    # category data element should be formatted as follows: [data_type, category_type]
    # warning: some category names overlap (find a way to avoid any problems)
    category_dict = {
        'hitter1': ['AVG', 'G', 'PA', 'AB', 'R', 'H', '2B', '3B', 'HR', 'TB', 'RBI', 'SAC', 'SF'],
        'hitter2': ['AVG', 'BB', 'IBB', 'HBP', 'SO', 'GDP', 'SLG', 'OBP', 'OPS', 'MH', 'RISP', 'PH-BA'],
        'pitcher1': ['ERA', 'G', 'W', 'L', 'SV', 'HLD', 'WPCT', 'IP', 'H', 'HR', 'BB', 'HBP', 'SO', 'R', 'ER', 'WHIP'],
        'pitcher2': ['ERA', 'CG', 'SHO', 'QS', 'BSV', 'TBF', 'NP', 'AVG', '2B', '3B', 'SAC', 'SF', 'IBB', 'WP', 'BK'],
        'defense': ['G', 'E', 'PKO', 'PO', 'A', 'DP', 'FPCT', 'PB', 'SB', 'CS', 'CS%'],
        'runner': ['G', 'SBA', 'SB', 'CS', 'SB%', 'OOB', 'PKO']
    }
    category_element_list = []
    for key in category_dict:
        for elem in category_dict[key]:
            category_element_list.append([key, elem])

    rank_total = []
    data_total = [[] for _ in range(len(categories))]

    for year in years:
        # read rank data
        rank_temp = []
        # give the location of the file
        loc = "/Users/joshuakim/PycharmProjects/KBO_Analysis/kbo_team_rank_data_" + str(year) + ".xlsx"

        # open Workbook
        workbook = xlrd.open_workbook(loc)
        worksheet = workbook.sheet_by_index(0)
        num_row_temp = worksheet.nrows - 1
        for i in range(num_row_temp):
            # multiplied 100 to obtain winning percentages (remove it to obtain pure decimal values)
            rank_temp.append([worksheet.cell_value(i + 1, 1), 100*worksheet.cell_value(i + 1, 6)])

        # appending data to augmented list in alphabetical order
        sorted_rank_temp = sort_team_data(rank_temp)
        for i in range(len(sorted_rank_temp)):
            rank_total.append(sorted_rank_temp[i][1])

        for j in range(len(categories)):
            category = categories[j]
            # read specified data
            data_temp = []
            # give the location of the file
            loc = "/Users/joshuakim/PycharmProjects/KBO_Analysis/kbo_team_"+str(category[0])+"_data_"+str(year)+"_regular_season.xlsx"

            # open Workbook
            workbook = xlrd.open_workbook(loc)
            worksheet = workbook.sheet_by_index(0)
            num_row_temp = worksheet.nrows - 2
            for i in range(num_row_temp):
                data_temp.append([worksheet.cell_value(i + 1, 1), worksheet.cell_value(i + 1, 2 + category_dict[str(category[0])].index(str(category[1])))])

            # appending data to augmented list in alphabetical order
            sorted_data_temp = sort_team_data(data_temp)
            for i in range(len(sorted_data_temp)):
                if str(category[1]) == 'IP':
                    data_total[j].append(convert_ip_to_float(sorted_data_temp[i][1]))
                    continue
                data_total[j].append(sorted_data_temp[i][1])
    return rank_total, data_total


if __name__ == '__main__':
    # years = [2001+i for i in range(18)]
    # categories = [['runner', 'SBA'], ['runner', 'CS'], ['runner', 'SB']]
    # coeffs, correl = analyze_team_data(years, categories)
    # print(years, categories)
    # print("Coefficients of Linear Fit: ", coeffs)
    # print("Multiple Correlation: ", correl)
    
    # years = [2001 + i for i in range(18)]
    # categories = [['hitter1', 'HR'], ['hitter1', '3B']]
    # print(years, categories)
    # coeffs, correl = compare_two_categories(years, categories)
    # print("Coefficients of Linear Fit: ", coeffs)
    # print("Pearson Correlation: ", correl)

    print(analyze_rank_data_wrt_time(9))
